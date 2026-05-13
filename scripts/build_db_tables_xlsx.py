#!/usr/bin/env python3
"""Generate database tables overview XLSX for robbo_personal_account_backend."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "database_tables_lk_overview.xlsx"

ROWS = [
    # table_name, go_model, database, status, purpose_ru
    (
        "— (отдельной таблицы users нет)",
        "UserDB (embedded)",
        "основная",
        "справочно",
        "Базовые поля пользователя встроены в таблицы ролей (student_dbs, teacher_dbs и т.д.); отдельной таблицы users в миграциях нет.",
    ),
    (
        "student_dbs",
        "StudentDB",
        "основная",
        "используется",
        "Учётные записи студентов: email, пароль, ФИО, роль, привязка к группе и филиалу (robbo_group_id, robbo_unit_id).",
    ),
    (
        "teacher_dbs",
        "TeacherDB",
        "основная",
        "используется",
        "Учётные записи преподавателей.",
    ),
    (
        "parent_dbs",
        "ParentDB",
        "основная",
        "используется",
        "Учётные записи родителей.",
    ),
    (
        "unit_admin_dbs",
        "UnitAdminDB",
        "основная",
        "используется",
        "Учётные записи администраторов филиала (unit admin).",
    ),
    (
        "super_admin_dbs",
        "SuperAdminDB",
        "основная",
        "используется",
        "Учётные записи суперадминистраторов.",
    ),
    (
        "free_listener_dbs",
        "FreeListenerDB",
        "основная",
        "используется",
        "Учётные записи внешних слушателей (free listener).",
    ),
    (
        "robbo_unit_dbs",
        "RobboUnitDB",
        "основная",
        "используется",
        "Филиалы Robbo (название, город и др.).",
    ),
    (
        "robbo_group_dbs",
        "RobboGroupDB",
        "основная",
        "используется",
        "Учебные группы (классы) внутри филиала.",
    ),
    (
        "children_of_parent_dbs",
        "ChildrenOfParentDB",
        "основная",
        "используется",
        "Связь many-to-many родитель ↔ ребёнок (student).",
    ),
    (
        "students_of_teacher_dbs",
        "StudentsOfTeacherDB",
        "основная",
        "используется",
        "Связь преподаватель ↔ студент (наставничество).",
    ),
    (
        "teachers_robbo_groups_dbs",
        "TeachersRobboGroupsDB",
        "основная",
        "используется",
        "Связь преподаватель ↔ учебная группа.",
    ),
    (
        "unit_admins_robbo_units_dbs",
        "UnitAdminsRobboUnitsDB",
        "основная",
        "используется",
        "Связь администратор филиала ↔ филиал.",
    ),
    (
        "course_dbs",
        "CourseDB",
        "основная",
        "используется",
        "Курсы (метаданные, синхронизация с EdX и внутренние поля).",
    ),
    (
        "course_relation_dbs",
        "CourseRelationDB",
        "основная",
        "используется",
        "Привязка курсов к сущностям (филиал, группа, студент, преподаватель, unit admin) по параметру object_id.",
    ),
    (
        "course_packet_dbs",
        "CoursePacketDB",
        "основная",
        "используется",
        "Пакеты курсов (группировка курсов).",
    ),
    (
        "course_api_media_collection_dbs",
        "CourseApiMediaCollectionDB",
        "основная",
        "используется",
        "Корневая сущность медиаданных курса (связь с course_id).",
    ),
    (
        "absolute_media_dbs",
        "AbsoluteMediaDB",
        "основная",
        "используется",
        "Абсолютные URL медиа, привязанные к коллекции курса.",
    ),
    (
        "media_dbs",
        "MediaDB",
        "основная",
        "используется",
        "Записи медиа (не изображения) для курса.",
    ),
    (
        "image_dbs",
        "ImageDB",
        "основная",
        "используется",
        "Изображения курса (превью и т.п.).",
    ),
    (
        "project_dbs",
        "ProjectDB",
        "основная",
        "только AutoMigrate; код не использует",
        "Устаревшая схема проектов Scratch (json в БД). Текущий код проектов работает через scratch_projects. Таблица создаётся миграцией GORM, но gateway не обращается — кандидат на удаление после проверки данных.",
    ),
    (
        "project_page_dbs",
        "ProjectPageDB",
        "основная",
        "только AutoMigrate; код не использует",
        "Устаревшие страницы проектов, связанные с project_dbs. Не используются текущими gateway — кандидат на удаление после миграции/бэкапа.",
    ),
    (
        "scratch_projects",
        "ScratchProjectDB",
        "проекты (projectsPostgres или fallback на основную)",
        "используется",
        "Актуальное хранилище проектов Scratch 3: заголовок, инструкция, scratch_vm_json, публичность, счётчик версий, soft delete.",
    ),
    (
        "scratch_project_versions",
        "ScratchProjectVersionDB",
        "проекты (тот же DSN, что и scratch_projects)",
        "используется",
        "История версий .sb3 (поле archive BYTEA) для проекта.",
    ),
    (
        "scratch_project_legacy_map",
        "ScratchProjectLegacyMapDB",
        "проекты",
        "используется",
        "Сопоставление старых идентификаторов проекта/страницы с UUID в новом хранилище.",
    ),
    (
        "cohort_dbs (условное имя GORM)",
        "CohortDB",
        "—",
        "не используется как таблица ЛК",
        "Модель есть в коде, но не входит в AutoMigrate; gateway CreateCohort возвращает пустой результат. Работа с когортами идёт через EdX API (маршруты /cohort в server.go закомментированы). Отдельная таблица когорт в PostgreSQL приложением не ведётся.",
    ),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблицы"

    headers = (
        "Таблица PostgreSQL (имя по GORM)",
        "Модель Go",
        "Какая БД (DSN)",
        "Статус для текущего кода ЛК",
        "Назначение",
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ROWS:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = (44, 28, 36, 28, 70)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    info = wb.create_sheet("Как читать", 0)
    info.append(
        [
            "Источник анализа: репозиторий robbo_personal_account_backend "
            "(package/db_client/postgres.go AutoMigrate, gateway-ы, config.yml)."
        ]
    )
    info.append([])
    info.append(
        [
            "Имена вида student_dbs соответствуют стандартному неймингу GORM v1.23 "
            "для структур с суффиксом DB (snake_case + plural)."
        ]
    )
    info.append([])
    info.append(
        [
            "Основная БД: postgres.postgresDsn. "
            "Проекты Scratch: projectsPostgres.postgresDsn; если пусто — используется тот же DSN, что и основная."
        ]
    )
    info.append([])
    info.append(
        [
            "Таблицы scratch_* не добавлены в AutoMigrate в коде — схема в БД проектов "
            "должна существовать (создана вручную или внешней миграцией)."
        ]
    )
    for r in info.iter_rows():
        for c in r:
            c.alignment = Alignment(wrap_text=True)
    info.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
